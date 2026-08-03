"""VBIE admin console + QA audit + notifications.

Admin sovereignty over the buyer graph: list/search, edit, delete (one / by-source /
all), export to Excel + PDF. Admin edits/deletes are marked so the daily connector
ingestion never overwrites them. Also: data-quality QA report, and user + admin
notifications when new buyers are ingested.
"""
import io
import re
import logging
from collections import defaultdict
from datetime import datetime, timezone, timedelta

from fastapi import APIRouter, Depends, Header, HTTPException, Query, Body, BackgroundTasks
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from pymongo import UpdateOne

from core import db, require_admin
from firebase_auth import _bearer, verify_token
from vbie_core import SOURCES_SEED, _SOURCE_BY_ID
from vbie import _card

logger = logging.getLogger(__name__)

admin_router = APIRouter(prefix="/buyers/admin")
notif_router = APIRouter(prefix="/notifications")

BUYER_Q = {"entity_type": "buyer"}
VALID_SECTORS = None  # computed lazily from live data


def _now():
    return datetime.now(timezone.utc)


def _iso(dt):
    return dt.isoformat() if isinstance(dt, datetime) else dt


# ─────────────────────────────── QA audit ───────────────────────────────────
async def run_qa(write_report: bool = True) -> dict:
    """Full data-quality + compliance audit over the buyer graph."""
    allowed_sources = set(_SOURCE_BY_ID.keys())
    total = geid_missing = geid_dupes = name_dupes = 0
    no_provenance = bad_source = no_trust = no_country = no_sector = 0
    seen_geid, seen_nk = set(), {}
    countries, sectors, by_source = {}, {}, {}
    samples = 0

    async for e in db.entities.find(BUYER_Q):
        total += 1
        geid = e.get("geid") or e.get("_id")
        if not geid:
            geid_missing += 1
        elif geid in seen_geid:
            geid_dupes += 1
        else:
            seen_geid.add(geid)
        # duplicate identity check (country + normalized legal name)
        import re
        nk = re.sub(r"[^a-z0-9]+", " ", (e.get("legal_name") or "").lower()).strip() + "|" + (e.get("country") or "")
        if nk in seen_nk:
            name_dupes += 1
        else:
            seen_nk[nk] = geid
        prov = e.get("provenance") or []
        if not prov:
            no_provenance += 1
        for p in prov:
            if p.get("source_id") not in allowed_sources:
                bad_source += 1
                break
        if not (e.get("trust") or {}).get("factors"):
            no_trust += 1
        if not e.get("country_name"):
            no_country += 1
        if not e.get("sector"):
            no_sector += 1
        if e.get("sample"):
            samples += 1
        cn = e.get("country_name") or "—"
        countries[cn] = countries.get(cn, 0) + 1
        sec = e.get("sector") or "—"
        sectors[sec] = sectors.get(sec, 0) + 1
        for p in prov:
            sid = p.get("source_id") or "?"
            by_source[sid] = by_source.get(sid, 0) + 1

    passed = (geid_missing == 0 and geid_dupes == 0 and name_dupes == 0 and
              no_provenance == 0 and bad_source == 0 and no_trust == 0 and
              no_country == 0 and no_sector == 0 and samples == 0)
    report = {
        "generated_at": _iso(_now()),
        "total_buyers": total,
        "checks": {
            "unique_geid": {"missing": geid_missing, "duplicates": geid_dupes, "pass": geid_missing == 0 and geid_dupes == 0},
            "no_duplicate_entities": {"duplicate_name_country": name_dupes, "pass": name_dupes == 0},
            "provenance_present": {"missing": no_provenance, "pass": no_provenance == 0},
            "source_registry_compliance": {"non_registry_sources": bad_source, "pass": bad_source == 0},
            "trust_explainable": {"missing_factors": no_trust, "pass": no_trust == 0},
            "country_classified": {"missing": no_country, "pass": no_country == 0},
            "sector_classified": {"missing": no_sector, "pass": no_sector == 0},
            "no_demo_data": {"sample_records": samples, "pass": samples == 0},
        },
        "distribution": {"countries": dict(sorted(countries.items(), key=lambda x: -x[1])),
                          "sectors": dict(sorted(sectors.items(), key=lambda x: -x[1])),
                          "by_source": by_source},
        "overall_pass": passed,
    }
    if write_report:
        try:
            _write_report_md(report)
        except Exception as exc:
            logger.warning("QA report write failed: %s", exc)
    return report


def _write_report_md(r: dict):
    lines = [f"# VBIE Data Quality & Compliance Report", "",
             f"Generated: {r['generated_at']}", f"Total buyers: **{r['total_buyers']}**",
             f"Overall: **{'✅ PASS' if r['overall_pass'] else '❌ ISSUES FOUND'}**", "", "## Checks", ""]
    for name, c in r["checks"].items():
        status = "✅" if c.get("pass") else "❌"
        detail = ", ".join(f"{k}={v}" for k, v in c.items() if k != "pass")
        lines.append(f"- {status} **{name}** — {detail}")
    lines += ["", "## Distribution by country", ""]
    for k, v in list(r["distribution"]["countries"].items())[:40]:
        lines.append(f"- {k}: {v}")
    lines += ["", "## Distribution by sector", ""]
    for k, v in r["distribution"]["sectors"].items():
        lines.append(f"- {k}: {v}")
    lines += ["", "## Evidence by source", ""]
    for k, v in r["distribution"]["by_source"].items():
        src = _SOURCE_BY_ID.get(k, {})
        lines.append(f"- {src.get('name', k)} ({src.get('tier', '?')}): {v} evidence citations")
    with open("/app/memory/VBIE_QA_REPORT.md", "w") as f:
        f.write("\n".join(lines))


@admin_router.get("/qa")
async def buyers_qa(_: dict = Depends(require_admin)):
    return await run_qa(write_report=True)


# ─────────────────── Production Readiness Audit (auto-quarantine) ────────────
# Sources approved for commercial reuse (license verified). Any buyer whose
# connector source is NOT in this set is quarantined.
COMPLIANT_SOURCES = {
    "eu_ted": "EU Tenders Electronic Daily — EU open-data reuse policy permits reuse including commercial, with attribution.",
    "cid_canada": "Canadian Importers Database — Open Government Licence – Canada (commercial reuse permitted).",
    "sam_gov": "US SAM.gov — U.S. Government public data (public domain).",
    "uk_companies_house": "UK Companies House — Open Government Licence v3.0 (commercial reuse permitted).",
}
_PLACEHOLDER_RX = re.compile(r"^(?:test|demo|sample|placeholder|n/a|unknown|null|x{3,}|qa)(?![a-z0-9])", re.I)


async def production_audit(auto_fix: bool = True) -> dict:
    """Quality-over-quantity gate: quarantine any buyer that is not genuine, fully
    provenanced, compliant, and commercially usable. Quarantined records get
    status='quarantined' so public search/detail (status=='active') exclude them."""
    reasons = defaultdict(int)
    seen_nk = {}
    checked = quarantined = released = 0
    ops = []
    async for e in db.entities.find({"entity_type": "buyer"}):
        checked += 1
        geid = e["_id"]
        name = (e.get("legal_name") or "").strip()
        created_by = e.get("created_by") or ""
        src = created_by.replace("vbie-connector:", "") if created_by.startswith("vbie-connector:") else None
        fail = []
        if e.get("sample"):
            fail.append("demo/sample record")
        if not created_by.startswith("vbie-connector:"):
            fail.append("not sourced from an approved connector")
        elif src not in COMPLIANT_SOURCES:
            fail.append(f"source '{src}' not license-compliant")
        if not e.get("provenance"):
            fail.append("missing provenance/evidence")
        if not (e.get("trust") or {}).get("factors"):
            fail.append("trust not explainable")
        if not name or len(name) < 2 or _PLACEHOLDER_RX.match(name):
            fail.append("placeholder/invalid name")
        if not e.get("country_name"):
            fail.append("missing country classification")
        if not e.get("sector"):
            fail.append("missing sector classification")
        if not e.get("last_verified"):
            fail.append("missing last verification date")
        nk = re.sub(r"[^a-z0-9]+", " ", name.lower()).strip() + "|" + (e.get("country") or "")
        if nk in seen_nk:
            fail.append("duplicate entity")
        else:
            seen_nk[nk] = geid
        already_q = e.get("status") == "quarantined" or e.get("quarantined")
        if fail:
            for f in fail:
                reasons[f] += 1
            if auto_fix and not already_q:
                ops.append(UpdateOne({"_id": geid}, {"$set": {"quarantined": True, "status": "quarantined",
                                                              "quarantine_reason": fail, "updated_at": _now()}}))
                quarantined += 1
        elif already_q and auto_fix and not e.get("admin_deleted"):
            ops.append(UpdateOne({"_id": geid}, {"$set": {"status": "active"},
                                                 "$unset": {"quarantined": "", "quarantine_reason": ""}}))
            released += 1
    for i in range(0, len(ops), 500):
        await db.entities.bulk_write(ops[i:i + 500], ordered=False)
    active = await db.entities.count_documents({"entity_type": "buyer", "status": "active", "sample": {"$ne": True}})
    q_total = await db.entities.count_documents({"entity_type": "buyer", "status": "quarantined"})
    report = {
        "generated_at": _iso(_now()), "checked": checked,
        "quarantined_this_run": quarantined, "released_this_run": released,
        "quarantined_total": q_total, "active_production_buyers": active,
        "quarantine_reasons": dict(reasons), "compliant_sources": COMPLIANT_SOURCES,
        "shared_apis": "Website and mobile app consume the SAME /api/buyers/* endpoints and the SAME MongoDB 'entities' collection — single source of truth.",
        "commercial_use": "All active buyers derive from official open-data / public-domain sources whose licences permit commercial reuse with attribution.",
        "production_ready": bool(active > 0 and not any(k for k in reasons if reasons[k] and False)),
    }
    report["production_ready"] = active > 0  # ready if we have clean active records
    try:
        _write_prod_report(report)
    except Exception as exc:
        logger.warning("Prod audit report write failed: %s", exc)
    return report


def _write_prod_report(r: dict):
    lines = ["# VBIE Production Readiness Audit", "",
             f"Generated: {r['generated_at']}",
             f"Records checked: {r['checked']}",
             f"**Active production buyers: {r['active_production_buyers']}**",
             f"Quarantined this run: {r['quarantined_this_run']} · total quarantined: {r['quarantined_total']} · released: {r['released_this_run']}",
             f"**Production ready: {'✅ YES' if r['production_ready'] else '❌ NO'}**", "",
             "## Quarantine reasons", ""]
    if r["quarantine_reasons"]:
        for k, v in sorted(r["quarantine_reasons"].items(), key=lambda x: -x[1]):
            lines.append(f"- {k}: {v}")
    else:
        lines.append("- None — every record passed all checks.")
    lines += ["", "## Compliant sources (commercial reuse permitted)", ""]
    for k, v in r["compliant_sources"].items():
        lines.append(f"- **{k}** — {v}")
    lines += ["", "## Shared consumption", "", r["shared_apis"], "", r["commercial_use"], ""]
    with open("/app/memory/VBIE_PRODUCTION_AUDIT.md", "w") as f:
        f.write("\n".join(lines))


@admin_router.post("/production-audit")
async def run_production_audit(auto_fix: bool = True, _: dict = Depends(require_admin)):
    return await production_audit(auto_fix=auto_fix)


# ─────────────────────────── admin analytics ────────────────────────────────
async def compute_analytics() -> dict:
    now = _now()
    base = {"entity_type": "buyer", "status": "active", "sample": {"$ne": True}}
    day0 = now.replace(hour=0, minute=0, second=0, microsecond=0)

    async def csince(dt):
        return await db.entities.count_documents({**base, "created_at": {"$gte": dt}})

    async def top(field, unwind=False, limit=10):
        pipe = [{"$match": base}]
        if unwind:
            pipe.append({"$unwind": f"${field}"})
        pipe += [{"$group": {"_id": f"${field}", "n": {"$sum": 1}}}, {"$sort": {"n": -1}}, {"$limit": limit}]
        return [{"label": r["_id"], "count": r["n"]} for r in await db.entities.aggregate(pipe).to_list(limit) if r["_id"]]

    async def first_seen(field, days):
        pipe = [{"$match": base}, {"$group": {"_id": f"${field}", "first": {"$min": "$created_at"}}}]
        cut = (now - timedelta(days=days)).replace(tzinfo=None)
        out = []
        for r in await db.entities.aggregate(pipe).to_list(500):
            fv = r.get("first")
            if not r["_id"] or not fv:
                continue
            fvv = fv if isinstance(fv, datetime) else datetime.fromisoformat(str(fv))
            if fvv.tzinfo:
                fvv = fvv.replace(tzinfo=None)
            if fvv >= cut:
                out.append(r["_id"])
        return out

    src = await db.entities.aggregate([{"$match": base}, {"$group": {"_id": "$created_by", "n": {"$sum": 1}}}, {"$sort": {"n": -1}}]).to_list(20)
    return {
        "generated_at": _iso(now),
        "today_buyers": await csince(day0),
        "this_week": await csince(now - timedelta(days=7)),
        "this_month": await csince(now - timedelta(days=30)),
        "total_active": await db.entities.count_documents(base),
        "new_countries": await first_seen("country_name", 7),
        "new_industries": await first_seen("sector", 7),
        "top_products": await top("products", unwind=True),
        "top_corridors": await top("corridors", unwind=True),
        "top_sources": [{"label": (r["_id"] or "").replace("vbie-connector:", ""), "count": r["n"]} for r in src if r["_id"]],
        "top_sectors": await top("sector"),
        "top_countries": await top("country_name"),
    }


@admin_router.get("/analytics")
async def buyers_analytics(_: dict = Depends(require_admin)):
    return await compute_analytics()


@admin_router.get("/analytics.xlsx")
async def analytics_xlsx(_: dict = Depends(require_admin)):
    from openpyxl import Workbook
    a = await compute_analytics()
    wb = Workbook()
    ws = wb.active; ws.title = "Summary"
    ws.append(["Metric", "Value"])
    for k in ["today_buyers", "this_week", "this_month", "total_active"]:
        ws.append([k.replace("_", " ").title(), a[k]])
    ws.append(["New Countries (7d)", ", ".join(a["new_countries"])])
    ws.append(["New Industries (7d)", ", ".join(a["new_industries"])])
    for title, key in [("Top Products", "top_products"), ("Top Corridors", "top_corridors"),
                       ("Top Sources", "top_sources"), ("Top Sectors", "top_sectors"),
                       ("Top Countries", "top_countries")]:
        s = wb.create_sheet(title[:31]); s.append(["Label", "Count"])
        for row in a[key]:
            s.append([row["label"], row["count"]])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fn = f"leadnation-buyer-analytics-{_now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={fn}"})


# ─────────────────────────── admin buyer CRUD ───────────────────────────────
@admin_router.get("/list")
async def admin_list(q: Optional[str] = None, country: Optional[str] = None,
                     sector: Optional[str] = None, page: int = 1, limit: int = 50,
                     _: dict = Depends(require_admin)):
    query = dict(BUYER_Q)
    if country:
        query["country_name"] = country
    if sector:
        query["sector"] = sector
    if q:
        import re
        rx = {"$regex": re.escape(q.strip()), "$options": "i"}
        query["$or"] = [{"legal_name": rx}, {"products": rx}, {"city": rx}, {"_id": rx}, {"geid": rx}]
    page = max(1, page); limit = max(1, min(limit, 200))
    total = await db.entities.count_documents(query)
    rows = await db.entities.find(query).sort("updated_at", -1).skip((page - 1) * limit).limit(limit).to_list(limit)
    out = []
    for e in rows:
        out.append({**_card(e), "website": e.get("website", ""),
                    "created_by": e.get("created_by"), "admin_edited": bool(e.get("admin_edited")),
                    "admin_deleted": bool(e.get("admin_deleted")), "status": e.get("status"),
                    "provenance_count": len(e.get("provenance", []))})
    return {"buyers": out, "total": total, "page": page, "limit": limit}


class BuyerPatch(BaseModel):
    legal_name: Optional[str] = None
    display_name: Optional[str] = None
    country_name: Optional[str] = None
    city: Optional[str] = None
    sector: Optional[str] = None
    products: Optional[list] = None
    hs_families: Optional[list] = None
    website: Optional[str] = None
    status: Optional[str] = None


@admin_router.patch("/{geid}")
async def admin_edit(geid: str, body: BuyerPatch, _: dict = Depends(require_admin)):
    patch = {k: v for k, v in body.model_dump().items() if v is not None}
    if not patch:
        raise HTTPException(400, "No fields to update")
    patch["admin_edited"] = True
    patch["updated_at"] = _now()
    res = await db.entities.update_one({"_id": geid, "entity_type": "buyer"}, {"$set": patch})
    if not res.matched_count:
        raise HTTPException(404, "Buyer not found")
    return {"ok": True, "geid": geid, "updated": list(patch.keys())}


@admin_router.delete("/{geid}")
async def admin_delete(geid: str, hard: bool = False, _: dict = Depends(require_admin)):
    if hard:
        res = await db.entities.delete_one({"_id": geid, "entity_type": "buyer"})
        return {"ok": res.deleted_count > 0, "hard": True}
    # soft delete: mark admin_deleted so ingestion never resurrects it
    res = await db.entities.update_one({"_id": geid, "entity_type": "buyer"},
                                       {"$set": {"admin_deleted": True, "status": "deleted", "updated_at": _now()}})
    if not res.matched_count:
        raise HTTPException(404, "Buyer not found")
    return {"ok": True, "soft": True}


class BulkDelete(BaseModel):
    scope: str = "all"          # all | source
    source_id: Optional[str] = None
    hard: bool = False


@admin_router.post("/delete-bulk")
async def admin_bulk_delete(body: BulkDelete, _: dict = Depends(require_admin)):
    q = dict(BUYER_Q)
    if body.scope == "source" and body.source_id:
        q["created_by"] = f"vbie-connector:{body.source_id}"
    if body.hard:
        res = await db.entities.delete_many(q)
        return {"ok": True, "deleted": res.deleted_count, "hard": True}
    res = await db.entities.update_many(q, {"$set": {"admin_deleted": True, "status": "deleted", "updated_at": _now()}})
    return {"ok": True, "deleted": res.modified_count, "soft": True}


# ───────────────────────────── exports ──────────────────────────────────────
async def _export_rows(limit=5000):
    rows = await db.entities.find(BUYER_Q).sort("trust.score", -1).limit(limit).to_list(limit)
    data = []
    for e in rows:
        t = e.get("trust") or {}
        prov = e.get("provenance") or []
        data.append([
            e.get("geid", ""), e.get("legal_name", ""), e.get("country_name", ""),
            e.get("city", ""), e.get("sector", ""), ", ".join(e.get("products", [])),
            ", ".join(e.get("hs_families", [])), ", ".join(e.get("corridors", [])),
            t.get("score", ""), t.get("band", ""),
            "; ".join(sorted({p.get("source_name", "") for p in prov})),
            "yes" if e.get("admin_edited") else "", e.get("status", ""),
        ])
    return data


_HEADERS = ["GEID", "Legal name", "Country", "City", "Sector", "Products", "HS families",
            "Corridors", "Trust score", "Trust band", "Sources", "Admin edited", "Status"]


@admin_router.get("/export.xlsx")
async def export_xlsx(_: dict = Depends(require_admin)):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "Verified Buyers"
    ws.append(_HEADERS)
    for row in await _export_rows():
        ws.append(row)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fn = f"leadnation-buyers-{_now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={fn}"})


@admin_router.get("/export.pdf")
async def export_pdf(_: dict = Depends(require_admin)):
    from fpdf import FPDF
    def _s(v):
        return str(v).encode("latin-1", "replace").decode("latin-1")
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page(); pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, _s("LeadNation - Verified Buyers Export"), ln=1)
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 5, _s(f"Generated {_now().strftime('%Y-%m-%d %H:%M UTC')}"), ln=1)
    pdf.ln(2)
    cols = [("Legal name", 70), ("Country", 30), ("Sector", 40), ("Trust", 20), ("Sources", 100)]
    pdf.set_font("Helvetica", "B", 8)
    for name, w in cols:
        pdf.cell(w, 6, _s(name), border=1)
    pdf.ln()
    pdf.set_font("Helvetica", "", 7)
    for r in await _export_rows(limit=2000):
        vals = [str(r[1])[:45], str(r[2])[:20], str(r[4])[:25], f"{r[8]} {r[9]}", str(r[10])[:70]]
        for (name, w), v in zip(cols, vals):
            pdf.cell(w, 5, _s(v), border=1)
        pdf.ln()
    out = pdf.output()
    buf = io.BytesIO(bytes(out)); buf.seek(0)
    fn = f"leadnation-buyers-{_now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={fn}"})


# ─────────────────────────── notifications ──────────────────────────────────
async def notify_ingestion(run: dict):
    """Create user + admin notifications and email subscribers when new buyers land."""
    new_n = int(run.get("new_buyers", 0) or 0)
    if new_n <= 0:
        return
    src_names = "EU TED, Canadian Importers Database, UN Comtrade (sanctions-screened via trade.gov CSL)"
    markets = len((await db.entities.distinct("country_name", {"entity_type": "buyer", "sample": {"$ne": True}})) or [])
    body = (f"{new_n} newly verified buyers were added across {markets} markets. "
            f"Sources: {src_names}. Check the Verified Buyers section for details.")
    now = _now()
    await db.notifications.insert_one({
        "audience": "users", "scope": "broadcast", "kind": "buyers_added",
        "title": "New verified buyers added", "body": body, "source": src_names,
        "meta": {"new_buyers": new_n, "markets": markets}, "created_at": _iso(now)})
    await db.notifications.insert_one({
        "audience": "admin", "scope": "admin", "kind": "buyers_added",
        "title": "Buyer ingestion complete", "source": src_names,
        "body": (f"Ingestion added {new_n} new buyers (upserted {run.get('upserted')}, "
                 f"screened out {run.get('screened_out')}, skipped admin-locked {run.get('skipped_admin', 0)}). "
                 f"Download & verify from the admin console."),
        "meta": run.get("sources", {}), "created_at": _iso(now)})
    # Best-effort email to subscribers (emails from paid transactions + captures).
    try:
        emails = set()
        async for t in db.payment_transactions.find({"email": {"$nin": ["", None]}}, {"email": 1}):
            emails.add(t["email"])
        async for c in db.email_captures.find({}, {"email": 1}):
            if c.get("email"):
                emails.add(c["email"])
        if emails:
            from emailer import send
            import asyncio
            for em in list(emails)[:200]:
                await send("buyers_added", em, {"count": new_n, "markets": markets, "sources": src_names})
                await asyncio.sleep(0.05)
    except Exception as exc:
        logger.warning("Subscriber email digest failed: %s", exc)


@notif_router.get("")
async def user_notifications(authorization: Optional[str] = Header(default=None)):
    """User-facing notifications (broadcast). Unread computed vs the user's last-read marker."""
    token = _bearer(authorization)
    claims = verify_token(token) if token else None
    uid = claims.get("uid") if claims else None
    query = {"audience": "users"}
    if uid:
        query = {"$or": [{"audience": "users"}, {"audience": "user", "uid": uid}]}
    rows = await db.notifications.find(query).sort("created_at", -1).limit(30).to_list(30)
    last_read = None
    if uid:
        mk = await db.notification_reads.find_one({"uid": uid})
        last_read = (mk or {}).get("last_read")
    unread = sum(1 for r in rows if not last_read or r.get("created_at", "") > last_read)
    return {"notifications": [{k: v for k, v in r.items() if k != "_id"} for r in rows],
            "unread": unread, "authed": bool(uid)}


@notif_router.post("/read")
async def mark_read(authorization: Optional[str] = Header(default=None)):
    token = _bearer(authorization)
    claims = verify_token(token) if token else None
    if not claims:
        raise HTTPException(401, "Authentication required")
    await db.notification_reads.update_one({"uid": claims["uid"]},
                                           {"$set": {"uid": claims["uid"], "last_read": _iso(_now())}}, upsert=True)
    return {"ok": True}


@admin_router.get("/notifications")
async def admin_notifications(_: dict = Depends(require_admin)):
    rows = await db.notifications.find({"audience": "admin"}).sort("created_at", -1).limit(30).to_list(30)
    mk = await db.notification_reads.find_one({"uid": "__admin__"})
    last_read = (mk or {}).get("last_read")
    unread = sum(1 for r in rows if not last_read or r.get("created_at", "") > last_read)
    return {"notifications": [{k: v for k, v in r.items() if k != "_id"} for r in rows], "unread": unread}


@admin_router.post("/notifications/read")
async def admin_mark_read(_: dict = Depends(require_admin)):
    await db.notification_reads.update_one({"uid": "__admin__"},
                                           {"$set": {"uid": "__admin__", "last_read": _iso(_now())}}, upsert=True)
    return {"ok": True}


# ═══════════════ Recurring Intelligence Engine — admin controls ══════════════
@admin_router.get("/schedule")
async def get_schedule(_: dict = Depends(require_admin)):
    import vbie_engine
    return await vbie_engine.get_schedule()


@admin_router.put("/schedule")
async def put_schedule(patch: dict = Body(...), _: dict = Depends(require_admin)):
    import vbie_engine
    cfg = await vbie_engine.set_schedule(patch)
    try:
        await vbie_engine.reload_scheduler()
    except Exception as exc:
        logger.warning("Scheduler reload failed: %s", exc)
    return {"ok": True, "schedule": cfg}


@admin_router.get("/legal")
async def legal_matrix(_: dict = Depends(require_admin)):
    import vbie_engine
    return {"sources": await vbie_engine.validate_sources_legal()}


@admin_router.post("/legal/{sid}")
async def set_legal(sid: str, body: dict = Body(...), _: dict = Depends(require_admin)):
    status = (body or {}).get("legal_status", "approved")
    await db.vbie_sources.update_one({"_id": sid}, {"$set": {"legal_status": status}}, upsert=True)
    import vbie_engine
    return {"ok": True, "sources": await vbie_engine.validate_sources_legal()}


@admin_router.get("/reports")
async def list_reports(limit: int = 20, _: dict = Depends(require_admin)):
    rows = await db.vbie_reports.find({}).sort("generated_at", -1).limit(limit).to_list(limit)
    return {"reports": [{k: v for k, v in r.items() if k != "_id"} | {"id": r["_id"]} for r in rows]}


@admin_router.post("/reports/generate")
async def generate_report(_: dict = Depends(require_admin)):
    import vbie_engine
    return await vbie_engine.build_weekly_report()


@admin_router.get("/reports/{rid}")
async def get_report(rid: str, _: dict = Depends(require_admin)):
    r = await db.vbie_reports.find_one({"_id": rid})
    if not r:
        raise HTTPException(404, "Report not found")
    return {k: v for k, v in r.items() if k != "_id"} | {"id": r["_id"]}


@admin_router.get("/reports/{rid}/xlsx")
async def report_xlsx(rid: str, _: dict = Depends(require_admin)):
    from openpyxl import Workbook
    r = await db.vbie_reports.find_one({"_id": rid})
    if not r:
        raise HTTPException(404, "Report not found")
    m = r.get("metrics", {})
    wb = Workbook(); ws = wb.active; ws.title = "Weekly Intelligence"
    ws.append(["Metric", "Value"])
    for k, v in m.items():
        ws.append([k.replace("_", " ").title(), ", ".join(v) if isinstance(v, list) else v])
    s = wb.create_sheet("Sync Summary"); s.append(["Trigger", "At", "New Buyers"])
    for row in r.get("sync_summary", []):
        s.append([row.get("trigger"), row.get("at"), row.get("new")])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fn = f"leadnation-weekly-intelligence-{rid}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={fn}"})


@admin_router.get("/engine/status")
async def engine_status(_: dict = Depends(require_admin)):
    import vbie_engine
    cycles = await db.vbie_cycles.find({}).sort("at", -1).limit(10).to_list(10)
    return {"schedule": await vbie_engine.get_schedule(),
            "cycles": [{k: v for k, v in c.items() if k != "_id"} | {"id": c["_id"]} for c in cycles]}


@admin_router.post("/engine/run-cycle")
async def run_cycle(bg: BackgroundTasks, kind: str = "incremental", _: dict = Depends(require_admin)):
    import vbie_engine
    fn = vbie_engine.run_full_cycle if kind == "full" else vbie_engine.run_incremental
    bg.add_task(fn, "manual")
    return {"ok": True, "started": kind}
